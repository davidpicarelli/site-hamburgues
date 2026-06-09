"""Excel report generator for forensic loan audit."""
from __future__ import annotations

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="0B1D3A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
LABEL_FONT = Font(bold=True, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="0B1D3A", name="Calibri")
THIN = Side(border_style="thin", color="D7DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = 'R$ #,##0.00;[Red]-R$ #,##0.00'
PCT_FMT = '0.0000%'


def _write_schedule_sheet(ws, title: str, rows: List[dict], show_obs: bool = True):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    headers = ["Nº", "Data", "Dias", "Saldo Inicial", "Juros", "Amortização", "Parcela", "Saldo Final", "Observações"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r["n"]).alignment = Alignment(horizontal="center")
        # Date
        from datetime import date as _date
        try:
            y, m, dday = map(int, r["due_date"].split("-"))
            ws.cell(row=i, column=2, value=f"{dday:02d}/{m:02d}/{y}").alignment = Alignment(horizontal="center")
        except Exception:
            ws.cell(row=i, column=2, value=str(r["due_date"]))
        ws.cell(row=i, column=3, value=r["days"]).alignment = Alignment(horizontal="center")
        for col, key in [(4, "opening_balance"), (5, "interest"), (6, "amortization"), (7, "installment"), (8, "closing_balance")]:
            cc = ws.cell(row=i, column=col, value=r[key])
            cc.number_format = MONEY_FMT
            cc.alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=9, value=r.get("note", "") if show_obs else "")
        for col in range(1, 10):
            ws.cell(row=i, column=col).border = BORDER
    total_row = 4 + len(rows)
    ws.cell(row=total_row, column=1, value="TOTAIS").font = LABEL_FONT
    total_int = sum(r["interest"] for r in rows)
    total_amort = sum(r["amortization"] for r in rows)
    total_inst = sum(r["installment"] for r in rows)
    for col, val in [(5, total_int), (6, total_amort), (7, total_inst)]:
        c = ws.cell(row=total_row, column=col, value=val)
        c.number_format = MONEY_FMT
        c.font = LABEL_FONT
        c.fill = PatternFill("solid", fgColor="F1F3F6")
        c.alignment = Alignment(horizontal="right")
    for col in (1, 5, 6, 7):
        ws.cell(row=total_row, column=col).border = BORDER
    widths = [6, 14, 8, 18, 16, 18, 16, 18, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_excel(result: dict) -> bytes:
    wb = Workbook()

    contract = result["contract"]
    bacen = result.get("bacen") or {}
    indicators = result.get("indicators", {})
    assumptions = result.get("assumptions", {})

    # ---- Sheet 1: Dados do contrato ----
    ws = wb.active
    ws.title = "Dados do Contrato"
    ws["A1"] = "LAUDO PERICIAL — RECONSTITUIÇÃO DE CONTRATO DE FINANCIAMENTO"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = "Método Hamburguês (juros simples sobre saldo devedor) vs. Tabela Price/SAC"
    ws["A2"].font = Font(italic=True, color="475569")
    ws.merge_cells("A2:D2")

    iof = contract.get("iof", 0)
    fees = contract.get("fees", 0)
    insurance = contract.get("insurance", 0)
    gross = contract["principal"] + iof + fees + insurance
    rows_input = [
        ("Valor financiado (principal líquido)", contract["principal"], MONEY_FMT),
        ("IOF embutido", iof, MONEY_FMT),
        ("Tarifas embutidas", fees, MONEY_FMT),
        ("Seguros embutidos", insurance, MONEY_FMT),
        ("Total bruto financiado", gross, MONEY_FMT),
        ("Taxa contratada (mensal)", contract.get("monthly_rate", 0) or 0, PCT_FMT),
        ("Taxa contratada (anual equivalente)", contract.get("annual_rate", 0) or 0, PCT_FMT),
        ("Prazo (meses)", contract["term_months"], None),
        ("Data de início", contract["start_date"], None),
        ("Data 1ª parcela", contract.get("first_due_date", ""), None),
        ("Sistema do contrato (original)", contract.get("original_system", "price").upper(), None),
        ("Parcela contratada (informada)", contract.get("contracted_installment") or 0, MONEY_FMT),
    ]
    for i, (k, v, fmt) in enumerate(rows_input, start=4):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=v)
        if fmt:
            c.number_format = fmt
        for col in (1, 2):
            ws.cell(row=i, column=col).border = BORDER

    # BACEN block
    base = 4 + len(rows_input) + 1
    ws.cell(row=base, column=1, value="REFERÊNCIA BACEN (SGS)").font = Font(bold=True, color="0B1D3A", size=12)
    bacen_rows = [
        ("Série", f"{bacen.get('series_code','')} — {bacen.get('series_label','')}"),
        ("Unidade", bacen.get("unit", "") or ""),
        ("Período", f"{bacen.get('start_date','')} → {bacen.get('end_date','')}"),
        ("Média observada", bacen.get("fetched_average") or bacen.get("manual_rate") or 0),
        ("Modo", bacen.get("mode", "manual")),
    ]
    for i, (k, v) in enumerate(bacen_rows, start=base + 1):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        ws.cell(row=i, column=2, value=v)
        for col in (1, 2):
            ws.cell(row=i, column=col).border = BORDER

    # Indicators block
    base2 = base + len(bacen_rows) + 2
    ws.cell(row=base2, column=1, value="INDICADORES").font = Font(bold=True, color="0B1D3A", size=12)
    ind_rows = [
        ("Excesso cobrado (Price − Hamburguês) em juros", indicators.get("excess_interest_vs_hamburgues", 0), MONEY_FMT),
        ("Taxa contratada acima da média BACEN?", "SIM" if indicators.get("above_bacen") else "NÃO", None),
        ("Indícios de anatocismo?", "SIM" if indicators.get("anatocism") else "NÃO", None),
    ]
    for i, (k, v, fmt) in enumerate(ind_rows, start=base2 + 1):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=v)
        if fmt:
            c.number_format = fmt
        for col in (1, 2):
            ws.cell(row=i, column=col).border = BORDER

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 32

    # ---- Sheets: schedules ----
    s_price = wb.create_sheet("Tabela Price")
    _write_schedule_sheet(s_price, "TABELA PRICE — Contrato Original (juros compostos, parcela fixa)", result["price"]["rows"], show_obs=False)
    s_sac = wb.create_sheet("SAC")
    _write_schedule_sheet(s_sac, "SISTEMA SAC — Amortização Constante (juros compostos)", result["sac"]["rows"], show_obs=False)
    s_h = wb.create_sheet("Método Hamburguês")
    _write_schedule_sheet(s_h, "MÉTODO HAMBURGUÊS — Recálculo com pagamentos reais (juros simples, sem anatocismo)", result["hamburgues"]["rows"])

    # ---- Sheet: Comparative ----
    s = wb.create_sheet("Comparativo")
    s["A1"] = "RESUMO COMPARATIVO"
    s["A1"].font = TITLE_FONT
    s.merge_cells("A1:E1")
    headers = ["Método", "Total Juros (R$)", "Total Pago (R$)", "Amortizado (R$)", "Saldo Final (R$)"]
    for col, h in enumerate(headers, start=1):
        c = s.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    for i, key in enumerate(["price", "sac", "hamburgues"], start=4):
        m = result[key]
        label = {"price": "Tabela Price (Original)", "sac": "SAC", "hamburgues": "Método Hamburguês"}[key]
        s.cell(row=i, column=1, value=label).font = LABEL_FONT
        for col, val in zip([2, 3, 4, 5], [m["total_interest"], m["total_paid"], m["total_amortization"], m["final_balance"]]):
            c = s.cell(row=i, column=col, value=val)
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal="right")
        for col in range(1, 6):
            s.cell(row=i, column=col).border = BORDER
    # diff row
    diff_int = result["price"]["total_interest"] - result["hamburgues"]["total_interest"]
    diff_paid = result["price"]["total_paid"] - result["hamburgues"]["total_paid"]
    s.cell(row=7, column=1, value="Diferença (Price − Hamburguês)").font = Font(bold=True, color="B42318")
    s.cell(row=7, column=2, value=diff_int).number_format = MONEY_FMT
    s.cell(row=7, column=3, value=diff_paid).number_format = MONEY_FMT
    s.cell(row=7, column=2).font = Font(bold=True, color="B42318")
    s.cell(row=7, column=3).font = Font(bold=True, color="B42318")
    for col in range(1, 6):
        s.cell(row=7, column=col).border = BORDER
        s.cell(row=7, column=col).fill = PatternFill("solid", fgColor="FEE4E2")
    widths = [35, 22, 22, 22, 22]
    for idx, w in enumerate(widths, start=1):
        s.column_dimensions[get_column_letter(idx)].width = w

    # ---- Sheet: Premissas ----
    s = wb.create_sheet("Premissas")
    s["A1"] = "PREMISSAS E METODOLOGIA"
    s["A1"].font = TITLE_FONT
    s.merge_cells("A1:B1")
    items = [
        ("Convenção de contagem de dias", "30 (taxa diária = taxa mensal / 30)" if assumptions.get("day_count") == "30" else "365 (taxa diária = taxa anual / 365)"),
        ("Arredondamento", "2 casas decimais (centavos), método half-up"),
        ("Capitalização no Método Hamburguês", "NÃO há — juros simples; juros não pagos não são incorporados ao saldo devedor"),
        ("Alocação do pagamento", "1º quita juros acumulados; 2º amortiza principal"),
        ("Tabela Price", "Parcela fixa, juros compostos sobre saldo devedor"),
        ("SAC", "Amortização constante, juros compostos sobre saldo devedor"),
        ("Fonte da taxa de mercado", "BACEN — Sistema Gerenciador de Séries Temporais (SGS)"),
    ]
    for i, (k, v) in enumerate(items, start=3):
        ws.cell(row=i, column=1, value=k)
        s.cell(row=i, column=1, value=k).font = LABEL_FONT
        s.cell(row=i, column=2, value=v)
        for col in (1, 2):
            s.cell(row=i, column=col).border = BORDER
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
